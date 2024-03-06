#merge cells 
import openpyxl
from openpyxl.styles import Font

wb=openpyxl.Workbook()
ws=wb.active

ws.merge_cells('A1:F2')
ws['A1']='MERGED CELLS '

merged_cell=ws['A1']
font=Font(size=16,color='FF0000')
merged_cell.font=font

wb.save('example2.xlsx')

#Red: (255, 0, 0) corresponds to FF0000 in hexadecimal.
#Green: (0, 255, 0) corresponds to 00FF00 in hexadecimal.
#Blue: (0, 0, 255) corresponds to 0000FF in hexadecimal.