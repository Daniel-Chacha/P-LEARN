
import pandas as pd 
import matplotlib.pyplot as plt 
import shutil
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment

def edit_individual_results():
    stud_fno=int(input("Enter the student's file number: "))
    df1=pd.read_excel('a-results2.xlsx')
    df1.index+=1
    column_name='FILE_NO.'
    num=len(df1)


    matching_row=df1[df1[column_name]==stud_fno]
    if matching_row.empty:
        print(f"There is no match of the File Number {stud_fno}")
        edit_individual_results()
    else:
        z=matching_row.index[0]
        #print(z)
        print(df1.loc[[z]])

        def confirm_edit():
            k=input('Enter the subject to change: ')
            intended_subject=k.lower()
            if intended_subject == 'mathematics':
                edit_subject='MATH'
            elif intended_subject == 'english':
                edit_subject= 'ENG'
            elif intended_subject=='kiswahili':
                edit_subject='KISW'
            elif intended_subject=='chemistry':
                edit_subject='CHEM'
            elif intended_subject=='biology':
                edit_subject='BIO'
            elif intended_subject=='physics':
                edit_subject='PHYC'
            elif intended_subject=='geography':
                edit_subject='GEO'
            elif intended_subject=='computer':
                edit_subject='COMP'
            else:
                print('ERROR!! No existing subject name')
                confirm_edit()

            g=int(input('Enter the new score: '))
            #df1.set_index('FILE_NO.',inplace=True)
            df1.at[z,edit_subject]=g

            df1['TOTAL']=df1[['MATH','ENG','KISW','CHEM','BIO','PHYC','GEO','COMP']].sum(axis=1)
            #print(df1)

            def tot_avg_grade():

                def grading(z):
                    if z>=90:
                        gr='A'
                    elif z>=80:
                        gr='A-'
                    elif z>=75:
                        gr='B+'
                    elif z>=70:
                        gr='B'
                    elif z>=65:
                        gr='B-'
                    elif z>=60:
                        gr='C+'
                    elif z>=55:
                        gr='C'
                    elif z>=50:
                        gr='C-'
                    elif z>=45:
                        gr='D+'
                    elif z>=40:
                        gr='D'
                    elif z>=35:
                        gr='D-'
                    else:
                        gr='E'
                    return gr

                file_no=list(df1['FILE_NO.'])
                names=list(df1['NAMES'])
                math=list(df1['MATH'])
                eng=list(df1['ENG'])
                kisw=list(df1['KISW'])
                chem=list(df1['CHEM'])
                bio=list(df1['BIO'])
                phyc=list(df1['PHYC'])
                geo=list(df1['GEO'])
                comp=list(df1['COMP'])
                total=list(df1['TOTAL'])
                average=[]
                grade=[]
                subject_total=[]
                subject_avg=[]
                subject_grade=[]

                for x in range(len(df1)):
                    avg=total[x]/8
                    average.append(avg)
                    w=grading(avg)
                    grade.append(w)

                subjects=[math,eng,kisw,chem,bio,phyc,geo,comp,total,average]
                for y in subjects:
                    grandtotal=sum(y)
                    subject_total.append(grandtotal)
                    s_avg= grandtotal/len(df1)
                    subject_avg.append(s_avg)
                    s_grade=grading(s_avg)
                    if y is total:
                        s_grade='-'
                    subject_grade.append(s_grade)

                data={
                    'FILE_NO.' :file_no,
                    'NAMES'    :names,
                    'MATH'     :math,
                    'ENG'      :eng,
                    'KISW'     :kisw,
                    'CHEM'     :chem,
                    'BIO'      :bio,
                    'PHYC'     :phyc,
                    'GEO'      :geo,
                    'COMP'     :comp,
                    'TOTAL'    :total,
                    'AVERAGE'  :average,
                    'GRADE'    :grade,
                    

                }
                df2=pd.DataFrame(data)
                df2.index+=1
                #print(df2)

                def excel_file(df):
                    sorted=df.sort_values(by='TOTAL',ascending=False)
                    sorted['RANK']=sorted['TOTAL'].rank(ascending=False,method='dense')

                    #a file and pass the dataframe to it
                    with pd.ExcelWriter('a-results2.xlsx',engine='openpyxl',mode='w')as file:
                        sorted.to_excel(file,index=False,startrow=0,startcol=0,sheet_name='sheet1')
                    
                    grandtotals=sorted.iloc[:,2:12].sum()
                    sorted.loc[len(sorted)+num]=['SubjectTotal','-',grandtotals[0],grandtotals[1],grandtotals[2],grandtotals[3],grandtotals[4],grandtotals[5],grandtotals[6],grandtotals[7],grandtotals[8],grandtotals[9],'-','-']

                    newgrand_avg=['SubjectAverage','-',subject_avg[0],subject_avg[1],subject_avg[2],subject_avg[3],subject_avg[4],subject_avg[5],subject_avg[6],subject_avg[7],subject_avg[8],subject_avg[9],'-']
                    extra2=pd.DataFrame([newgrand_avg],columns=df.columns)
                    sorted=sorted._append(extra2, ignore_index=True)

                    newgrand_grade=['SubjectGrades','-',subject_grade[0],subject_grade[1],subject_grade[2],subject_grade[3],subject_grade[4],subject_grade[5],subject_grade[6],subject_grade[7],subject_grade[8],subject_grade[9],'-']
                    extra3=pd.DataFrame([newgrand_grade],columns=df.columns)
                    sorted=sorted._append(extra3, ignore_index=True)
                    
                    grands=grandtotals[0:8]      
                    ranked_subjects=grands.rank(ascending=False,method='min')
                    sorted.loc[len(sorted)]=['SubjectRank','-',ranked_subjects[0],ranked_subjects[1],ranked_subjects[2],ranked_subjects[3],ranked_subjects[4],ranked_subjects[5],ranked_subjects[6],ranked_subjects[7],'-','-','-','-']

                    sorted.index+=1
                    #print(sorted)

                    #creating an excel workbook and passing the dataframe to it

                    with pd.ExcelWriter('a-results3.xlsx',engine='openpyxl',mode='w') as writer:
                            sorted.to_excel(writer,index=False,startrow=4,startcol=0,sheet_name='Sheet1')

                    #call the file and append additional text above the table

                    wb=openpyxl.load_workbook('a-results3.xlsx')
                    ws=wb.active

                    c=[['A1','N1'],['A2','N2'],['A3','N3'],['A4','N4']]
                    s=['HIDDENVILLE  HIGHSCHOOL','FORM 3 NORTH','TERM  2','']
                    for start_cell,end_cell in c:
                        ws.merge_cells(f'{start_cell}:{end_cell}')
                        cell=ws[start_cell]
                        cell.alignment=Alignment(horizontal='center',vertical='center')
                    
                    ws['A1'].value='HIDDENVILLE  HIGHSCHOOL'
                    ws['A1'].font=Font(size=21)
                    ws['A2'].value='FORM 3 NORTH'
                    ws['A2'].font=Font(size=17)
                    ws['A3'].value='TERM  2'
                    ws['A3'].font=Font(size=17)
                    ws['A4'].value='MOCK  EXAM'
                    ws['A4'].font=Font(size=17)

                    wb.save('a-results3.xlsx')
                
                excel_file(df2)
            tot_avg_grade()
        confirm_edit()

def edit_whole_class():
    print("subject selection available include:")
    print(' Mathematics , English ,Kiswahili, Chemistry, Biology, Physics, Geography, Computer')
    k=input('Enter your choice: ')
    sub_choice=k.lower()
    if sub_choice == 'mathematics':
        search_column='MATH'
    elif sub_choice == 'english':
        search_column='ENG'
    elif sub_choice == 'kiswahili':
        search_column='KISW'
    elif sub_choice == 'chemistry':
        search_column= 'CHEM'
    elif sub_choice == 'biology':
        search_column='BIO'
    elif sub_choice == 'physics':
        search_column='PHYC'
    elif sub_choice == 'geography':
        search_column='GEO'
    elif sub_choice == 'computer':
        search_column='COMP'
    else:
        print('ERROR!! Wrong Entry, Re-enter your choice')
        edit_whole_class()
    df1=pd.read_excel('a-results2.xlsx')
    df1.index+=1
    num=len(df1)
    selected_columns=df1[['FILE_NO.','NAMES',search_column]]
    print(selected_columns)

    for y in range(len(df1)):
        y+=1
        print(selected_columns.loc[[y]])
        df1.at[y,search_column]=int(input('Enter the new score: '))
    
    df1['TOTAL']=df1[['MATH','ENG','KISW','CHEM','BIO','PHYC','GEO','COMP']].sum(axis=1)
    #print(df1)
 
    def tot_avg_grade():

        def grading(z):
            if z>=90:
                gr='A'
            elif z>=80:
                gr='A-'
            elif z>=75:
                gr='B+'
            elif z>=70:
                gr='B'
            elif z>=65:
                gr='B-'
            elif z>=60:
                gr='C+'
            elif z>=55:
                gr='C'
            elif z>=50:
                gr='C-'
            elif z>=45:
                gr='D+'
            elif z>=40:
                gr='D'
            elif z>=35:
                gr='D-'
            else:
                gr='E'
            return gr

        file_no=list(df1['FILE_NO.'])
        names=list(df1['NAMES'])
        math=list(df1['MATH'])
        eng=list(df1['ENG'])
        kisw=list(df1['KISW'])
        chem=list(df1['CHEM'])
        bio=list(df1['BIO'])
        phyc=list(df1['PHYC'])
        geo=list(df1['GEO'])
        comp=list(df1['COMP'])
        total=list(df1['TOTAL'])
        average=[]
        grade=[]
        subject_total=[]
        subject_avg=[]
        subject_grade=[]

        for x in range(len(df1)):
            avg=total[x]/8
            average.append(avg)
            w=grading(avg)
            grade.append(w)

        subjects=[math,eng,kisw,chem,bio,phyc,geo,comp,total,average]
        for y in subjects:
            grandtotal=sum(y)
            subject_total.append(grandtotal)
            s_avg= grandtotal/len(df1)
            subject_avg.append(s_avg)
            s_grade=grading(s_avg)
            if y is total:
                s_grade='-'
            subject_grade.append(s_grade)

        data={
            'FILE_NO.' :file_no,
            'NAMES'    :names,
            'MATH'     :math,
            'ENG'      :eng,
            'KISW'     :kisw,
            'CHEM'     :chem,
            'BIO'      :bio,
            'PHYC'     :phyc,
            'GEO'      :geo,
            'COMP'     :comp,
            'TOTAL'    :total,
            'AVERAGE'  :average,
            'GRADE'    :grade,
            

        }
        df2=pd.DataFrame(data)
        df2.index+=1
        #print(df2)

        def excel_file(df):
            sorted=df.sort_values(by='TOTAL',ascending=False)
            sorted['RANK']=sorted['TOTAL'].rank(ascending=False,method='dense')

            #a file and pass the dataframe to it
            with pd.ExcelWriter('a-results2.xlsx',engine='openpyxl',mode='w')as file:
                sorted.to_excel(file,index=False,startrow=0,startcol=0,sheet_name='sheet1')

            grandtotals=sorted.iloc[:,2:12].sum()
            sorted.loc[len(sorted)+num]=['SubjectTotal','-',grandtotals[0],grandtotals[1],grandtotals[2],grandtotals[3],grandtotals[4],grandtotals[5],grandtotals[6],grandtotals[7],grandtotals[8],grandtotals[9],'-','-']

            newgrand_avg=['SubjectAverage','-',subject_avg[0],subject_avg[1],subject_avg[2],subject_avg[3],subject_avg[4],subject_avg[5],subject_avg[6],subject_avg[7],subject_avg[8],subject_avg[9],'-']
            extra2=pd.DataFrame([newgrand_avg],columns=df.columns)
            sorted=sorted._append(extra2, ignore_index=True)

            newgrand_grade=['SubjectGrades','-',subject_grade[0],subject_grade[1],subject_grade[2],subject_grade[3],subject_grade[4],subject_grade[5],subject_grade[6],subject_grade[7],subject_grade[8],subject_grade[9],'-']
            extra3=pd.DataFrame([newgrand_grade],columns=df.columns)
            sorted=sorted._append(extra3, ignore_index=True)
            
            grands=grandtotals[0:8]
            ranked_subjects=grands.rank(ascending=False,method='min')
            sorted.loc[len(sorted)]=['SubjectRank','-',ranked_subjects[0],ranked_subjects[1],ranked_subjects[2],ranked_subjects[3],ranked_subjects[4],ranked_subjects[5],ranked_subjects[6],ranked_subjects[7],'-','-','-','-']
            
            sorted.index+=1
            #print(sorted)

            #creating an excel workbook and passing the dataframe to it

            with pd.ExcelWriter('a-results3.xlsx',engine='openpyxl',mode='w') as writer:
                    sorted.to_excel(writer,index=False,startrow=4,startcol=0,sheet_name='Sheet1')

            #call the file and append additional text above the table

            wb=openpyxl.load_workbook('a-results3.xlsx')
            ws=wb.active

            c=[['A1','N1'],['A2','N2'],['A3','N3'],['A4','N4']]
            s=['HIDDENVILLE  HIGHSCHOOL','FORM 3 NORTH','TERM  2','']
            for start_cell,end_cell in c:
                ws.merge_cells(f'{start_cell}:{end_cell}')
                cell=ws[start_cell]
                cell.alignment=Alignment(horizontal='center',vertical='center')
            
            ws['A1'].value='HIDDENVILLE  HIGHSCHOOL'
            ws['A1'].font=Font(size=21)
            ws['A2'].value='FORM 3 NORTH'
            ws['A2'].font=Font(size=17)
            ws['A3'].value='TERM  2'
            ws['A3'].font=Font(size=17)
            ws['A4'].value='MOCK  EXAM'
            ws['A4'].font=Font(size=17)

            wb.save('a-results3.xlsx')
        
        excel_file(df2)
    tot_avg_grade()

def edit():
    print('Edit the results of an individual student or for the whole class in a particular subject')
    decision=int(input("Enter '1' to edit for an individual student or '2' for the whole class: "))

    if decision==1:
        edit_individual_results()
    elif decision==2:
        edit_whole_class()
    else:
        print('ERROR!! Wrong choice entered')
        edit()

edit()