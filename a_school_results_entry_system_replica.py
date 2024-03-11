import pandas as pd 
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment

num=int(input('How many students: '))

file_no=[]
names=[]
math=[]
eng=[]
kisw=[]
chem=[]
bio=[]
phyc=[]
geo=[]
comp=[]
total=[]
average=[]
grade=[]
subject_total=[]
subject_avg=[]
subject_grade=[]

#assign grades
def grading(z):
    if z>=0 and z<=100:
        if z>=90:
            gr='A'
        elif z>=80:
            gr='A-'
        elif z>=75:
            gr='B+'
        elif z>=70:
            gr='B'
        elif z>=65:
            gr='B-'
        elif z>=60:
            gr='C+'
        elif z>=55:
            gr='C'
        elif z>=50:
            gr='C-'
        elif z>=45:
            gr='D+'
        elif z>=40:
            gr='D'
        elif z>=35:
            gr='D-'
        else:
            gr='E'
    else:
        gr='Wrong marks entered'
    return gr

for x in range(num):
    h=int(input('FILE.NO: '))

    while h in file_no:
        print('ERROR!! The FileNo already exists in another entry')
        print('Re-enter the FileNo')
        h=int(input('Enter FILENo. '))
    file_no.append(h)
    a=(input('NAME: '))
    names.append(a)

start_entry={
    'FILE_NO.' :file_no,
    'NAMES'  :names
}
df1=pd.DataFrame(start_entry)
df1.index+=1

def subject_choice():
    k=input('Enter name of subject to key in results: ')
    sub_choice=k.lower()
    if sub_choice == 'mathematics':
        sub_selected=math
    elif sub_choice == 'english':
        sub_selected=eng
    elif sub_choice == 'kiswahili':
        sub_selected=kisw
    elif sub_choice == 'chemistry':
        sub_selected= chem
    elif sub_choice == 'biology':
        sub_selected=bio
    elif sub_choice == 'physics':
        sub_selected=phyc
    elif sub_choice == 'geography':
        sub_selected=geo
    elif sub_choice == 'computer':
        sub_selected=comp
    else:
        print('ERROR!! Wrong Entry, Re-enter your choice')
        subject_choice()

   
    x=1
    while x<= len(df1):
        print(df1.loc[[x]])
        a=int(input(("Enter student's mark: ")))
        sub_selected.append(a)
        x+=1
    
    def another_subject():
        print('Do you want to enter marks for another subject? ')
        des=input('YES or NO: ')
        des1=des.lower()
        if des1 == 'yes':
            subject_choice()
        elif des1=='no':
            print(f'Marks entry for {sub_choice} completed.')
        else:
            print('Wrong choice entered!! Re-try')
            another_subject()
    another_subject()

subject_choice()

for e in range(num):
    t=0
    if math:
        t+=math[e]
    if eng:
        t+=eng[e]
    if kisw:
        t+=kisw[e]
    if chem:
        t+=chem[e]
    if bio:
        t+=bio[e]
    if phyc:
        t+=phyc[e]
    if geo:
        t+=geo[e]
    if comp:
        t+=comp[e]

    total.append(t)
    avg=t/8
    average.append(avg)
    w=grading(avg)
    grade.append(w)

data={
    'FILE_NO.' :file_no,
    'NAMES'    :names,
    'MATH'     :math,
    'ENG'      :eng,
    'KISW'     :kisw,
    'CHEM'     :chem,
    'BIO'      :bio,
    'PHYC'     :phyc,
    'GEO'      :geo,
    'COMP'     :comp,
    'TOTAL'    :total,
    'AVERAGE'  :average,
    'GRADE'    :grade
}

series_dict={key:pd.Series(value,dtype=object) for key, value in data.items()}

df2=pd.DataFrame(series_dict)
df2.index+=1
#print(df2)

#total marks,avg and grade of each subject
subjects=[math,eng,kisw,chem,bio,phyc,geo,comp,total,average]
for y in subjects:
    grandtotal=sum(y)
    subject_total.append(grandtotal)
    s_avg= grandtotal/num
    subject_avg.append(s_avg)
    s_grade=grading(s_avg)
    if y is total:
        s_grade='-'
    subject_grade.append(s_grade)



def excel_file(df):
    sorted=df.sort_values(by='TOTAL',ascending=False)
    sorted['RANK']=sorted['TOTAL'].rank(ascending=False,method='dense')

    #a file and pass the dataframe to it
    with pd.ExcelWriter('a-results2.xlsx',engine='openpyxl',mode='w')as file:
        sorted.to_excel(file,index=False,startrow=0,startcol=0,sheet_name='sheet1')

    #newsubtotal=['SubjectTotal','-',subject_total[0],subject_total[1],subject_total[2],subject_total[3],subject_total[4],subject_total[5],subject_total[6],subject_total[7],subject_total[8],subject_total[9],'-']
    #extra1=pd.DataFrame([newsubtotal],columns=df.columns)
    #sorted=sorted._append(extra1, ignore_index=True)
    grandtotals=sorted.iloc[:,2:12].sum()
    sorted.loc[len(sorted)+num]=['SubjectTotal','-',grandtotals[0],grandtotals[1],grandtotals[2],grandtotals[3],grandtotals[4],grandtotals[5],grandtotals[6],grandtotals[7],grandtotals[8],grandtotals[9],'-','-']
    #print(grandtotals) 

    newgrand_avg=['SubjectAverage','-',subject_avg[0],subject_avg[1],subject_avg[2],subject_avg[3],subject_avg[4],subject_avg[5],subject_avg[6],subject_avg[7],subject_avg[8],subject_avg[9],'-']
    extra2=pd.DataFrame([newgrand_avg],columns=df.columns)
    sorted=sorted._append(extra2, ignore_index=True)

    newgrand_grade=['SubjectGrades','-',subject_grade[0],subject_grade[1],subject_grade[2],subject_grade[3],subject_grade[4],subject_grade[5],subject_grade[6],subject_grade[7],subject_grade[8],subject_grade[9],'-']
    extra3=pd.DataFrame([newgrand_grade],columns=df.columns)
    sorted=sorted._append(extra3, ignore_index=True)
    
    grands=grandtotals[0:8]
    ranked_subjects=grands.rank(ascending=False,method='min')
    sorted.loc[len(sorted)]=['SubjectRank','-',ranked_subjects[0],ranked_subjects[1],ranked_subjects[2],ranked_subjects[3],ranked_subjects[4],ranked_subjects[5],ranked_subjects[6],ranked_subjects[7],'-','-','-','-']

    sorted.index+=1

    #creating an excel workbook and passing the dataframe to it

    with pd.ExcelWriter('a-results3.xlsx',engine='openpyxl',mode='w') as writer:
            sorted.to_excel(writer,index=False,startrow=4,startcol=0,sheet_name='Sheet1')

    #call the file and append additional text above the table

    wb=openpyxl.load_workbook('a-results3.xlsx')
    ws=wb.active

    c=[['A1','N1'],['A2','N2'],['A3','N3'],['A4','N4']]
    s=['HIDDENVILLE  HIGHSCHOOL','FORM 3 NORTH','TERM  2','']
    for start_cell,end_cell in c:
        ws.merge_cells(f'{start_cell}:{end_cell}')
        cell=ws[start_cell]
        cell.alignment=Alignment(horizontal='center',vertical='center')
    
    ws['A1'].value='HIDDENVILLE  HIGHSCHOOL'
    ws['A1'].font=Font(size=21)
    ws['A2'].value='FORM 3 NORTH'
    ws['A2'].font=Font(size=17)
    ws['A3'].value='TERM  2'
    ws['A3'].font=Font(size=17)
    ws['A4'].value='MOCK  EXAM'
    ws['A4'].font=Font(size=17)

    wb.save('a-results3.xlsx')
    
excel_file(df2)
