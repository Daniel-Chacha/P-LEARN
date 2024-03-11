import pandas as pd 
import matplotlib.pyplot as plt 
import shutil
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import sys

def studentresult():
    search_value=int(input('Enter the File Number of the Student: '))

    df1=pd.read_excel('a-results2.xlsx')
    #print(df)
    df1.index+=1
    column_name='FILE_NO.'


    matching_row=df1[df1[column_name]==search_value]
    if matching_row.empty:
        print(f"There is no match of the File Number {search_value}")
        studentresult()
    else:
        z=matching_row.index[0]
        #print(z)
        print(df1.loc[[z]])

    #print('Enter yes to delete or no to exit')
    def confirm_deletion():
        a=input("Enter 'yes' to delete or 'no' to exit: ")
        option=a.lower()
        #print(matching_row)
        if option =='yes':
            new_df=df1.drop(matching_row.index)
            num=len(new_df)
            #print(new_df)

            def grading(z):
                if z>=90:
                    gr='A'
                elif z>=80:
                    gr='A-'
                elif z>=75:
                    gr='B+'
                elif z>=70:
                    gr='B'
                elif z>=65:
                    gr='B-'
                elif z>=60:
                    gr='C+'
                elif z>=55:
                    gr='C'
                elif z>=50:
                    gr='C-'
                elif z>=45:
                    gr='D+'
                elif z>=40:
                    gr='D'
                elif z>=35:
                    gr='D-'
                else:
                    gr='E'
                return gr

            file_no=list(new_df['FILE_NO.'])
            names=list(new_df['NAMES'])
            math=list(new_df['MATH'])
            eng=list(new_df['ENG'])
            kisw=list(new_df['KISW'])
            chem=list(new_df['CHEM'])
            bio=list(new_df['BIO'])
            phyc=list(new_df['PHYC'])
            geo=list(new_df['GEO'])
            comp=list(new_df['COMP'])
            total=list(new_df['TOTAL'])
            average=list(new_df['AVERAGE'])
            grade=list(new_df['GRADE'])
            subject_total=[]
            subject_avg=[]
            subject_grade=[]

            subjects=[math,eng,kisw,chem,bio,phyc,geo,comp,total,average]
            for y in subjects:
                grandtotal=sum(y)
                subject_total.append(grandtotal)
                s_avg= grandtotal/num
                subject_avg.append(s_avg)
                s_grade=grading(s_avg)
                if y is total:
                    s_grade='-'
                subject_grade.append(s_grade)

            data={
                'FILE_NO.' :file_no,
                'NAMES'    :names,
                'MATH'     :math,
                'ENG'      :eng,
                'KISW'     :kisw,
                'CHEM'     :chem,
                'BIO'      :bio,
                'PHYC'     :phyc,
                'GEO'      :geo,
                'COMP'     :comp,
                'TOTAL'    :total,
                'AVERAGE'  :average,
                'GRADE'    :grade,
                

            }
            df2=pd.DataFrame(data)
            df2.index+=1
            #print(df2)

            def excel_file(df):
                sorted=df.sort_values(by='TOTAL',ascending=False)
                sorted['RANK']=sorted['TOTAL'].rank(ascending=False,method='dense')

                #a file and pass the dataframe to it
                with pd.ExcelWriter('a-results2.xlsx',engine='openpyxl',mode='w')as file:
                    sorted.to_excel(file,index=False,startrow=0,startcol=0,sheet_name='sheet1')
                
                grandtotals=sorted.iloc[:,2:12].sum()
                sorted.loc[len(sorted)+num]=['SubjectTotal','-',grandtotals[0],grandtotals[1],grandtotals[2],grandtotals[3],grandtotals[4],grandtotals[5],grandtotals[6],grandtotals[7],grandtotals[8],grandtotals[9],'-','-']


                newgrand_avg=['SubjectAverage','-',subject_avg[0],subject_avg[1],subject_avg[2],subject_avg[3],subject_avg[4],subject_avg[5],subject_avg[6],subject_avg[7],subject_avg[8],subject_avg[9],'-']
                extra2=pd.DataFrame([newgrand_avg],columns=df.columns)
                sorted=sorted._append(extra2, ignore_index=True)

                newgrand_grade=['SubjectGrades','-',subject_grade[0],subject_grade[1],subject_grade[2],subject_grade[3],subject_grade[4],subject_grade[5],subject_grade[6],subject_grade[7],subject_grade[8],subject_grade[9],'-']
                extra3=pd.DataFrame([newgrand_grade],columns=df.columns)
                sorted=sorted._append(extra3, ignore_index=True)

                grands=grandtotals[0:8]      
                ranked_subjects=grands.rank(ascending=False,method='min')
                sorted.loc[len(sorted)]=['SubjectRank','-',ranked_subjects[0],ranked_subjects[1],ranked_subjects[2],ranked_subjects[3],ranked_subjects[4],ranked_subjects[5],ranked_subjects[6],ranked_subjects[7],'-','-','-','-']

                sorted.index+=1
                #print(sorted)

                #creating an excel workbook and passing the dataframe to it

                with pd.ExcelWriter('a-results3.xlsx',engine='openpyxl',mode='w') as writer:
                        sorted.to_excel(writer,index=False,startrow=4,startcol=0,sheet_name='Sheet1')

                #call the file and append additional text above the table

                wb=openpyxl.load_workbook('a-results3.xlsx')
                ws=wb.active

                c=[['A1','N1'],['A2','N2'],['A3','N3'],['A4','N4']]
                s=['HIDDENVILLE  HIGHSCHOOL','FORM 3 NORTH','TERM  2','']
                for start_cell,end_cell in c:
                    ws.merge_cells(f'{start_cell}:{end_cell}')
                    cell=ws[start_cell]
                    cell.alignment=Alignment(horizontal='center',vertical='center')
                
                ws['A1'].value='HIDDENVILLE  HIGHSCHOOL'
                ws['A1'].font=Font(size=21)
                ws['A2'].value='FORM 3 NORTH'
                ws['A2'].font=Font(size=17)
                ws['A3'].value='TERM  2'
                ws['A3'].font=Font(size=17)
                ws['A4'].value='MOCK  EXAM'
                ws['A4'].font=Font(size=17)

                wb.save('a-results3.xlsx')
                
            excel_file(df2)
            #print(df2)

            #new_df.to_excel('a-results2.xlsx',index=False)
        elif option =='no':
            print('EXITING WITHOUT DELETING')
            sys.exit()
        else:
            print("Invalid Input: Please enter 'yes' or 'no' ")
            confirm_deletion()  
    confirm_deletion()
    

studentresult()
def repeat():
    print("Would you like to delete  another student's results?")
    decision=input(" 'YES' or 'NO': ")
    if decision=='yes':
        studentresult()
    elif decision=='no':
        sys.exit()
    else:
        print(' Wrong input!')
        repeat()

repeat()