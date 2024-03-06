#import matplotlib.pyplot as plt 
#import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment

new_df=pd.read_excel('a-results2.xlsx')
file_number=new_df['FILE_NO.']
new_file=[]
for x in file_number:
    new_file.append(x)

num=int(input('How many additional entries to enter: '))

file_no=[]
names=[]
math=[]
eng=[]
kisw=[]
chem=[]
bio=[]
phyc=[]
geo=[]
comp=[]
total=[]
average=[]
grade=[]
subject_total=[]
subject_avg=[]
subject_grade=[]

#assign grades
def grading(z):
    if z>=0 and z<=100:
        if z>=90:
            gr='A'
        elif z>=80:
            gr='A-'
        elif z>=75:
            gr='B+'
        elif z>=70:
            gr='B'
        elif z>=65:
            gr='B-'
        elif z>=60:
            gr='C+'
        elif z>=55:
            gr='C'
        elif z>=50:
            gr='C-'
        elif z>=45:
            gr='D+'
        elif z>=40:
            gr='D'
        elif z>=35:
            gr='D-'
        else:
            gr='E'
    else:
        gr='Wrong marks entered'
    return gr


print('Enter the scores for each student')
for x in range(num):
    h=int(input('FILE NUMBER: '))

    while h in new_file:
        print('ERROR!! The FileNo already exists in another entry')
        print('Re-enter the FileNo')
        h=int(input('Enter FILE NUMBER: '))
    while h in file_no:
        print('ERROR!! You already entered the file number')
        print('Re-enter the File Number')
        h=int(input('Enter FILE NUMBER: '))

    file_no.append(h)
    a=(input('NAME: '))
    names.append(a)
    b=int(input('MATH: '))
    math.append(b)
    c=int(input('ENG: '))
    eng.append(c)
    d=int(input('KISW: '))
    kisw.append(d)
    e=int(input('CHEM: '))
    chem.append(e)
    f=int(input('BIO: '))
    bio.append(f)
    g=int(input('PHYC: '))
    phyc.append(g)
    h=int(input('GEO: '))
    geo.append(h)
    i=int(input('COMP: '))
    comp.append(i)
    t=b+c+d+e+f+g+h+i
    total.append(t)
    avg=t/8
    average.append(avg)

    w=grading(avg)
    grade.append(w)

    if x<num:
        print('\nNew Entry\n')
    else:
        print('----')

#total marks,avg and grade of each subject
subjects=[math,eng,kisw,chem,bio,phyc,geo,comp,total,average]
for y in subjects:
    grandtotal=sum(y)
    subject_total.append(grandtotal)
    s_avg= grandtotal/num
    subject_avg.append(s_avg)
    s_grade=grading(s_avg)
    if y is total:
        s_grade='-'
    subject_grade.append(s_grade)

data={
    'FILE_NO.' :file_no,
    'NAMES'    :names,
    'MATH'     :math,
    'ENG'      :eng,
    'KISW'     :kisw,
    'CHEM'     :chem,
    'BIO'      :bio,
    'PHYC'     :phyc,
    'GEO'      :geo,
    'COMP'     :comp,
    'TOTAL'    :total,
    'AVERAGE'  :average,
    'GRADE'    :grade,
    

}
df=pd.DataFrame(data)
#df.index+=1

appended_df=pd.concat([new_df,df],ignore_index=True)
#print(df)

sorted=appended_df.sort_values(by='TOTAL',ascending=False)
sorted['RANK']=sorted['TOTAL'].rank(ascending=False,method='dense')

#a file and pass the dataframe to it
with pd.ExcelWriter('a-results2.xlsx',engine='openpyxl',mode='w')as file:
    sorted.to_excel(file,index=False,startrow=0,startcol=0,sheet_name='sheet1')

newsubtotal=['SubjectTotal','-',subject_total[0],subject_total[1],subject_total[2],subject_total[3],subject_total[4],subject_total[5],subject_total[6],subject_total[7],subject_total[8],subject_total[9],'-']
extra1=pd.DataFrame([newsubtotal],columns=df.columns)
sorted=sorted._append(extra1, ignore_index=True)

newgrand_avg=['SubjectAverage','-',subject_avg[0],subject_avg[1],subject_avg[2],subject_avg[3],subject_avg[4],subject_avg[5],subject_avg[6],subject_avg[7],subject_avg[8],subject_avg[9],'-']
extra2=pd.DataFrame([newgrand_avg],columns=df.columns)
sorted=sorted._append(extra2, ignore_index=True)

newgrand_grade=['SubjectGrades','-',subject_grade[0],subject_grade[1],subject_grade[2],subject_grade[3],subject_grade[4],subject_grade[5],subject_grade[6],subject_grade[7],subject_grade[8],subject_grade[9],'-']
extra3=pd.DataFrame([newgrand_grade],columns=df.columns)
sorted=sorted._append(extra3, ignore_index=True)

sorted.index+=1

#creating an excel workbook and passing the dataframe to it
with pd.ExcelWriter('a-results3.xlsx',engine='openpyxl',mode='w') as writer:
    sorted.to_excel(writer,index=False,startrow=4,startcol=0,sheet_name='Sheet1')

#call the file and append additional text above the table

wb=openpyxl.load_workbook('a-results3.xlsx')
ws=wb.active

c=[['A1','N1'],['A2','N2'],['A3','N3'],['A4','N4']]
s=['HIDDENVILLE  HIGHSCHOOL','FORM 3 NORTH','TERM  2','']
for start_cell,end_cell in c:
    ws.merge_cells(f'{start_cell}:{end_cell}')
    cell=ws[start_cell]
    cell.alignment=Alignment(horizontal='center',vertical='center')
    
ws['A1'].value='HIDDENVILLE  HIGHSCHOOL'
ws['A1'].font=Font(size=21)
ws['A2'].value='FORM 3 NORTH'
ws['A2'].font=Font(size=17)
ws['A3'].value='TERM  2'
ws['A3'].font=Font(size=17)
ws['A4'].value='MOCK  EXAM'
ws['A4'].font=Font(size=17)

wb.save('a-results3.xlsx')