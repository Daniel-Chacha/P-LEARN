import openpyxl

wb=openpyxl.Workbook()
sheet=wb.active

c1=sheet.cell(row=1,column=1)
c1.value='ANKIT'

c2=sheet.cell(row=1,column=2)
c2.value='RAI'

# Once have a Worksheet object, one can 
# access a cell object by its name also. 
# A2 means column = 1 & row = 2.
c3=sheet['A2']
c3.value='RAHUL'

c4=sheet['B2']
c4.value='RAI'

wb.save('demo.xlsx')